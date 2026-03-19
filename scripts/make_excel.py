#!/usr/bin/env python3
# =============================================================================
# make_excel.py — Genera Excel BOQ desde el JSON exportado por NavisBOQ
#
# Produce un archivo .xlsx con 3 hojas:
#   1. BOQ_Resumen   — Tabla ejecutiva con subtotales, lista para imprimir
#   2. Detalle_PowerBI — Tabla plana sin formato, ideal para conectar en Power BI
#   3. Pivot_Nivel   — Pivot rápido: Categoría × Nivel
#
# USO:
#   python make_excel.py C:\Users\TuUsuario\Desktop\BOQ_modelo_20250311.json
#   python make_excel.py archivo.json --output C:\Reportes\BOQ_Final.xlsx
#
# INSTALACIÓN DE DEPENDENCIA:
#   Abre PowerShell y ejecuta:
#   pip install openpyxl
# =============================================================================

import json
import sys
import os
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Falta la librería openpyxl.")
    print("Instálala ejecutando en PowerShell:")
    print("    pip install openpyxl")
    sys.exit(1)

# =============================================================================
# Colores corporativos — modifica según tu identidad visual
# =============================================================================
C_HEADER_BG  = "1F3864"   # Azul oscuro  → fondo de encabezados
C_HEADER_FG  = "FFFFFF"   # Blanco       → texto de encabezados
C_SUBTOTAL   = "2E75B6"   # Azul medio   → filas de subtotal
C_SUBTOTAL_FG= "FFFFFF"
C_TOTAL      = "1F3864"   # Azul oscuro  → fila gran total
C_TOTAL_FG   = "FFFFFF"
C_ALT        = "EBF3FB"   # Azul claro   → filas alternas
C_WARN       = "FFC000"   # Naranja      → elementos sin zona
C_ACCENT     = "00B0A0"   # Verde-teal   → acento inferior de header

THIN = Side(style="thin",   color="CCCCCC")
MED  = Side(style="medium", color="2E75B6")


def border_thin():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, value, width=None, ws=None, col=None):
    """Aplica estilo de encabezado a una celda."""
    cell.value     = value
    cell.font      = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=11)
    cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(
        left   = THIN,
        right  = THIN,
        top    = Side(style="medium", color=C_HEADER_BG),
        bottom = Side(style="medium", color=C_ACCENT))
    if width and ws and col:
        col_letter = get_column_letter(col)
        if ws.column_dimensions[col_letter].width < width:
            ws.column_dimensions[col_letter].width = width


# =============================================================================
# HOJA 1: BOQ_Resumen
# =============================================================================

def build_resumen(wb, resumen_rows, metadata):
    ws = wb.active
    ws.title = "BOQ_Resumen"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    archivo  = metadata.get("archivo", "")
    generado = metadata.get("generado", "")
    total    = metadata.get("total", 0)

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value     = f"BILL OF QUANTITIES  ·  {archivo.upper()}"
    cell.font      = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value     = f"Generado: {generado}   |   Total elementos: {total:,}"
    sub.font      = Font(name="Calibri", size=10, color="666666", italic=True)
    sub.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Encabezados ───────────────────────────────────────────────────────────
    headers = ["Edificio","Zona","Nivel","Categoría","Tipo","Cantidad","Unidad","# Elementos"]
    widths  = [16, 16, 14, 18, 30, 13, 9, 11]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        style_header(ws.cell(row=3, column=i), h, w, ws, i)
    ws.row_dimensions[3].height = 22
    ws.auto_filter.ref = "A3:H3"

    # ── Datos ─────────────────────────────────────────────────────────────────
    prev_cat  = None
    subtotals = {}   # cat → {"total": 0, "unidad": "", "n": 0}
    row_n     = 4

    for r in resumen_rows:
        cat = r.get("Categoria", "")

        # Insertar subtotal cuando cambia la categoría
        if prev_cat and prev_cat != cat:
            row_n = _write_subtotal(ws, row_n, prev_cat, subtotals[prev_cat])

        alt  = (row_n % 2 == 0)
        fill = PatternFill("solid", fgColor=C_ALT) if alt else None
        sin_zona = r.get("Zona", "") in ("Sin asignar", "")

        vals = [
            r.get("Edificio", ""),
            r.get("Zona", ""),
            r.get("Nivel", ""),
            cat,
            r.get("Tipo", ""),
            r.get("TotalCantidad", 0),
            r.get("Unidad", ""),
            r.get("NumElementos", 0),
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=row_n, column=ci, value=val)
            c.font      = Font(name="Calibri", size=10,
                               color=C_WARN if (sin_zona and ci == 2) else "000000")
            c.border    = border_thin()
            c.alignment = Alignment(
                horizontal = "right" if ci in (6, 8) else "left",
                vertical   = "center")
            if fill: c.fill = fill
            if ci == 6: c.number_format = "#,##0.00"
            if ci == 8: c.number_format = "#,##0"

        # Acumular subtotal
        if cat not in subtotals:
            subtotals[cat] = {"total": 0, "unidad": r.get("Unidad",""), "n": 0}
        subtotals[cat]["total"] += r.get("TotalCantidad", 0)
        subtotals[cat]["n"]     += r.get("NumElementos", 0)

        prev_cat = cat
        row_n   += 1

    # Último subtotal
    if prev_cat:
        row_n = _write_subtotal(ws, row_n, prev_cat, subtotals[prev_cat])

    # Gran total
    ws.merge_cells(f"A{row_n}:E{row_n}")
    gt = ws.cell(row=row_n, column=1, value="GRAN TOTAL")
    gt.font      = Font(name="Calibri", bold=True, size=11, color=C_TOTAL_FG)
    gt.fill      = PatternFill("solid", fgColor=C_TOTAL)
    gt.alignment = Alignment(horizontal="right", vertical="center")
    for ci in range(1, 6):
        c = ws.cell(row=row_n, column=ci)
        c.fill = PatternFill("solid", fgColor=C_TOTAL)
    total_n = ws.cell(row=row_n, column=8,
                      value=sum(v["n"] for v in subtotals.values()))
    total_n.font          = Font(name="Calibri", bold=True, size=11, color=C_TOTAL_FG)
    total_n.fill          = PatternFill("solid", fgColor=C_TOTAL)
    total_n.number_format = "#,##0"
    total_n.alignment     = Alignment(horizontal="right")
    ws.row_dimensions[row_n].height = 20


def _write_subtotal(ws, row_n, cat, data):
    ws.merge_cells(f"A{row_n}:E{row_n}")
    sc = ws.cell(row=row_n, column=1, value=f"Subtotal — {cat}")
    sc.font      = Font(name="Calibri", bold=True, size=10, color=C_SUBTOTAL_FG)
    sc.fill      = PatternFill("solid", fgColor=C_SUBTOTAL)
    sc.alignment = Alignment(horizontal="right", vertical="center")
    for ci in range(1, 6):
        ws.cell(row=row_n, column=ci).fill = PatternFill("solid", fgColor=C_SUBTOTAL)

    tv = ws.cell(row=row_n, column=6, value=round(data["total"], 2))
    tv.font          = Font(name="Calibri", bold=True, color=C_SUBTOTAL_FG)
    tv.fill          = PatternFill("solid", fgColor=C_SUBTOTAL)
    tv.number_format = "#,##0.00"
    tv.alignment     = Alignment(horizontal="right")

    uv = ws.cell(row=row_n, column=7, value=data["unidad"])
    uv.font = Font(name="Calibri", bold=True, color=C_SUBTOTAL_FG)
    uv.fill = PatternFill("solid", fgColor=C_SUBTOTAL)

    nv = ws.cell(row=row_n, column=8, value=data["n"])
    nv.font          = Font(name="Calibri", bold=True, color=C_SUBTOTAL_FG)
    nv.fill          = PatternFill("solid", fgColor=C_SUBTOTAL)
    nv.number_format = "#,##0"
    nv.alignment     = Alignment(horizontal="right")

    ws.row_dimensions[row_n].height = 18
    return row_n + 1


# =============================================================================
# HOJA 2: Detalle_PowerBI — tabla plana para Power BI
# =============================================================================

def build_detalle(wb, detalle_rows):
    ws = wb.create_sheet("Detalle_PowerBI")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["Proyecto","Edificio","Zona","Nivel",
               "Categoria","Tipo","Instancia","Cantidad","Unidad","ElementoId"]
    widths  = [20, 16, 16, 14, 18, 32, 28, 12, 8, 36]

    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        style_header(ws.cell(row=1, column=i), h, w, ws, i)
    ws.row_dimensions[1].height = 22

    for ri, row in enumerate(detalle_rows, start=2):
        alt  = (ri % 2 == 0)
        fill = PatternFill("solid", fgColor=C_ALT) if alt else None
        vals = [
            row.get("Proyecto",""),   row.get("Edificio",""),
            row.get("Zona",""),       row.get("Nivel",""),
            row.get("Categoria",""),  row.get("Tipo",""),
            row.get("Instancia",""),  row.get("Cantidad", 0),
            row.get("Unidad",""),     row.get("ElementoId","")
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Calibri", size=9)
            c.border    = border_thin()
            c.alignment = Alignment(
                vertical   = "center",
                horizontal = "right" if ci == 8 else "left")
            if fill: c.fill = fill
            if ci == 8: c.number_format = "#,##0.00"

    ws.auto_filter.ref = f"A1:J{ri}"


# =============================================================================
# HOJA 3: Pivot Categoría × Nivel
# =============================================================================

def build_pivot(wb, resumen_rows):
    ws = wb.create_sheet("Pivot_Cat_Nivel")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:Z1")
    t = ws["A1"]
    t.value     = "CANTIDADES POR CATEGORÍA × NIVEL"
    t.font      = Font(name="Calibri", bold=True, size=12, color=C_HEADER_FG)
    t.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    niveles  = sorted(set(r.get("Nivel","")    for r in resumen_rows))
    cats     = sorted(set(r.get("Categoria","") for r in resumen_rows))

    # Construir índice pivot
    idx = {}
    for r in resumen_rows:
        cat  = r.get("Categoria","")
        niv  = r.get("Nivel","")
        cant = r.get("TotalCantidad",0)
        uni  = r.get("Unidad","")
        if cat not in idx: idx[cat] = {}
        if niv not in idx[cat]: idx[cat][niv] = (0, uni)
        prev, u = idx[cat][niv]
        idx[cat][niv] = (prev + cant, u)

    # Encabezados
    style_header(ws.cell(row=2, column=1), "Categoría", 20, ws, 1)
    style_header(ws.cell(row=2, column=2), "Unidad",    9,  ws, 2)
    for i, niv in enumerate(niveles, start=3):
        style_header(ws.cell(row=2, column=i), niv, 12, ws, i)
    total_col = len(niveles) + 3
    style_header(ws.cell(row=2, column=total_col), "TOTAL", 13, ws, total_col)
    ws.row_dimensions[2].height = 20

    for ri, cat in enumerate(cats, start=3):
        alt  = (ri % 2 == 0)
        fill = PatternFill("solid", fgColor=C_ALT) if alt else None

        c1 = ws.cell(row=ri, column=1, value=cat)
        c1.font   = Font(name="Calibri", size=10, bold=True)
        c1.border = border_thin()
        if fill: c1.fill = fill

        units     = [v[1] for v in idx.get(cat,{}).values() if v[1]]
        uni_cat   = max(set(units), key=units.count) if units else ""
        c2        = ws.cell(row=ri, column=2, value=uni_cat)
        c2.font   = Font(name="Calibri", size=10)
        c2.border = border_thin()
        if fill: c2.fill = fill

        total_row = 0
        for i, niv in enumerate(niveles, start=3):
            val = idx.get(cat,{}).get(niv,(0,""))[0]
            cn  = ws.cell(row=ri, column=i, value=round(val,2) if val else None)
            cn.font          = Font(name="Calibri", size=10)
            cn.border        = border_thin()
            cn.alignment     = Alignment(horizontal="right")
            cn.number_format = "#,##0.00"
            if fill: cn.fill = fill
            total_row += val

        ct = ws.cell(row=ri, column=total_col, value=round(total_row,2))
        ct.font          = Font(name="Calibri", size=10, bold=True)
        ct.border        = border_thin()
        ct.alignment     = Alignment(horizontal="right")
        ct.number_format = "#,##0.00"
        ct.fill          = PatternFill("solid", fgColor="D9E2F3")


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Genera Excel BOQ desde JSON exportado por NavisBOQ plugin")
    parser.add_argument("json_file",
                        help="Ruta al archivo JSON generado por NavisBOQ (export_json)")
    parser.add_argument("--output", "-o",
                        help="Ruta del Excel de salida (opcional, por defecto junto al JSON)")
    args = parser.parse_args()

    if not os.path.exists(args.json_file):
        print(f"ERROR: No se encuentra el archivo: {args.json_file}")
        sys.exit(1)

    print(f"\n📂 Leyendo: {args.json_file}")
    with open(args.json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    metadata     = data.get("metadata", {})
    resumen_rows = data.get("resumen", [])
    detalle_rows = data.get("detalle", [])

    print(f"   → {len(resumen_rows):,} filas de resumen BOQ")
    print(f"   → {len(detalle_rows):,} filas de detalle (elementos)")

    wb = openpyxl.Workbook()

    print("📊 Generando hoja BOQ_Resumen...")
    build_resumen(wb, resumen_rows, metadata)

    print("📋 Generando hoja Detalle_PowerBI...")
    build_detalle(wb, detalle_rows)

    print("🔢 Generando hoja Pivot_Cat_Nivel...")
    build_pivot(wb, resumen_rows)

    # Determinar ruta de salida
    if args.output:
        out_path = args.output
    else:
        base     = os.path.splitext(args.json_file)[0]
        out_path = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    wb.save(out_path)

    print(f"\n✅ Excel generado exitosamente:")
    print(f"   {out_path}")
    print()
    print("📌 Cómo usar el Excel:")
    print("   • BOQ_Resumen    → Vista ejecutiva para cliente (imprimible)")
    print("   • Detalle_PowerBI → Conéctalo en Power BI como fuente de datos")
    print("   • Pivot_Cat_Nivel → Vista rápida Categoría × Nivel")
    print()


if __name__ == "__main__":
    main()
